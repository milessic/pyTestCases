import json
import sys
import os
import re
from pathlib import Path
from time import time
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QVBoxLayout,
    QHBoxLayout, QFrame, QFileDialog, QMessageBox, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QWidget, QTextEdit
)
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtGui import QIcon
from PyQt5 import QtGui

from src.models import TestCaseModel
from src.Elements import MyComboBox
from src.exports import Formatting


class PyTestCasesApp(QMainWindow):
    control_pressed = False
    shift_pressed = False
    def __init__(self, theme: str = None, actual_results_displayed=True, always_on_top: bool=True, column_width: int=200, 
                 preconditions_displayed: bool=True, default_exports_prefix: str="output_", column_names_row: int=3, 
                 test_case_data_starting_row: int=4, xlsx_test_cases_sheet_name: str="TEST CASES"):
        super().__init__()
        self.show_save_info = True
        self.test_cases = []
        self.dragging = False
        self.drag_position = QPoint()


        self.app_name = "PyTestCases"
        #self.xlsx_test_cases_sheet_name = xlsx_test_cases_sheet_name
        self.xlsx_test_cases_sheet_name = ""
        self.default_exports_prefix = default_exports_prefix
        self.editing_disabled = False
        self.column_names_row = column_names_row
        self.test_case_data_starting_row = test_case_data_starting_row
        self.actual_results_displayed = not actual_results_displayed
        self.preconditions_displayed = preconditions_displayed
        self.column_width = int(column_width)
        self.current_test_index = 0
        self.test_execution_id = None
        self.icon_path = os.path.join(os.path.dirname(__file__), "icon27.png")
        self.icon = QtGui.QIcon(self.icon_path)
        self.setWindowIcon(self.icon)

        self.initUI()
        self.setWindowFlags(Qt.FramelessWindowHint)
        if always_on_top:
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
        self.show()
        self.control_actual_results()

    def initUI(self):
        # Window title
        self.setWindowTitle(self.app_name)
        self.setGeometry(100, 100, 600, 400)

        # Main widget and layout
        self.main_widget = QWidget(self)
        self.setCentralWidget(self.main_widget)
        self.layout = QVBoxLayout(self.main_widget)

        # TopNav
        top_nav = QHBoxLayout()
        top_nav_l = QHBoxLayout()
        # left side of topnav
        self.top_nav_icon_pixmap = QtGui.QPixmap()
        self.top_nav_icon_pixmap.load(self.icon_path)
        self.top_nav_icon = QLabel()
        self.top_nav_icon.setPixmap(self.top_nav_icon_pixmap)
        self.top_nav_app_name = QLabel(self.app_name, self)

        top_nav_l.addWidget(self.top_nav_icon, alignment=Qt.AlignLeft)
        top_nav_l.addWidget(self.top_nav_app_name, alignment=Qt.AlignLeft)
        # right side of topnav 
        top_nav_r = QHBoxLayout()
        top_nav_r.setAlignment(Qt.AlignRight)
        self.minimize_button = QPushButton("-", self)
        self.minimize_button.setFixedWidth(30)
        self.minimize_button.clicked.connect(self.showMinimized)
        self.close_button = QPushButton("x", self)
        self.close_button.setFixedWidth(30)
        self.close_button.clicked.connect(sys.exit)
        top_nav_r.addWidget(self.minimize_button)
        top_nav_r.addWidget(self.close_button, alignment=Qt.AlignRight)
        top_nav.addLayout(top_nav_l)
        top_nav.addLayout(top_nav_r)
        self.layout.addLayout(top_nav)
        # Top Panel
        self.execution_id_label = QLabel("Test Execution ID:", self)
        self.execution_id_input = QLineEdit(self)
        self.load_tests_button = QPushButton("Load Tests", self)
        self.load_tests_button.clicked.connect(self.loadTests)
        self.export_combo = QComboBox(self)
        self.export_combo.addItems(["Export","Jira", "Xlsx"])
        self.export_combo.currentIndexChanged.connect(self.handleExport)

        top_layout = QHBoxLayout()
        top_layout.addWidget(self.execution_id_label)
        top_layout.addWidget(self.execution_id_input)
        top_layout.addWidget(self.load_tests_button)
        top_layout.addWidget(self.export_combo)
        self.layout.addLayout(top_layout)

        # Dropdown for test cases
        self.test_case_dropdown = QComboBox(self)
        self.test_case_dropdown.currentIndexChanged.connect(self.displayTestCase)
        self.test_case_dropdown.setMinimumWidth(240)

        # Test status display
        self.test_status_info_label = QLabel("Test Status:", self)
        self.test_status_label = QLabel("", self)
        dropdown_layout = QHBoxLayout()
        dropdown_layout.addWidget(self.test_case_dropdown)
        dropdown_layout.addWidget(self.test_status_info_label)
        dropdown_layout.addWidget(self.test_status_label)
        self.layout.addLayout(dropdown_layout)

        # Buttons for test status updates
        self.button_frame = QFrame(self)
        self.button_layout = QHBoxLayout(self.button_frame)

        self.pass_button = QPushButton("PASS", self.button_frame)
        self.fail_button = QPushButton("FAIL", self.button_frame)
        self.blocked_button = QPushButton("BLOCKED", self.button_frame)
        self.not_tested_button = QPushButton("NOT TESTED", self.button_frame)
        self.prev_button = QPushButton("Previous", self.button_frame)
        self.next_button = QPushButton("Next", self.button_frame)

        self.pass_button.clicked.connect(lambda: self.updateTestStatus("PASS"))
        self.fail_button.clicked.connect(lambda: self.updateTestStatus("FAIL"))
        self.blocked_button.clicked.connect(lambda: self.updateTestStatus("BLOCKED"))
        self.not_tested_button.clicked.connect(lambda: self.updateTestStatus("NOT TESTED"))
        self.prev_button.clicked.connect(self.previousTestCase)
        self.next_button.clicked.connect(self.nextTestCase)

        self.button_layout.addWidget(self.pass_button)
        self.button_layout.addWidget(self.fail_button)
        self.button_layout.addWidget(self.blocked_button)
        self.button_layout.addWidget(self.not_tested_button)
        self.button_layout.addWidget(self.prev_button)
        self.button_layout.addWidget(self.next_button)
        self.layout.addWidget(self.button_frame)

        # Preconditions display
        self.preconditions_text_frame = QFrame(self)
        self.preconditions_layout = QVBoxLayout(self.preconditions_text_frame)
        self.preconditions_text_layout = QHBoxLayout()
        self.preconditions_layout.addLayout(self.preconditions_text_layout)
        self.preconditions_label = QLabel("Preconditions", self.preconditions_text_frame)
        self.assignee_label = QLabel("Assignee", self.preconditions_text_frame)
        self.assignee_value = QLabel("", self.preconditions_text_frame)
        self.preconditions_text = QTextEdit("", self.preconditions_text_frame)
        #self.preconditions_text.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.preconditions_text_layout.addWidget(self.preconditions_label)
        self.preconditions_layout.addWidget(self.preconditions_text)
        self.preconditions_text_layout.addWidget(self.assignee_label)
        self.preconditions_text_layout.addWidget(self.assignee_value)
        self.layout.addWidget(self.preconditions_text_frame)

        # Test Case Table
        self.table_test_case_frame = QFrame(self)
        self.test_table_layout = QVBoxLayout(self.table_test_case_frame)
        self.test_case_table = QTableWidget(0, 3, self.table_test_case_frame)
        self.test_case_table.setHorizontalHeaderLabels(["Description", "Expected Result", "Actual Result"])
        self.test_case_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.test_case_table.horizontalHeader().setStretchLastSection(True)
        self.test_case_table.setWordWrap(True)
        self.test_case_table.setMinimumHeight(150) 
        # test case control button
        self.test_case_table_control_actual_resultss_btn = QPushButton("x", self.table_test_case_frame)
        self.test_case_table_control_actual_resultss_btn.clicked.connect(lambda: self.control_actual_results())
        self.test_case_table_control_actual_resultss_btn.setFixedSize(20, 20)

        self.test_table_layout.addWidget(self.test_case_table)
        self.layout.addWidget(self.table_test_case_frame)


    # events
    def keyReleaseEvent(self, e):
        if e.key() == Qt.Key_Control:
            self.control_pressed = False
        if e.key() == Qt.Key_Shift:
            self.shift_pressed = False

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Shift:
            self.shift_pressed = True
        if e.key() == Qt.Key_Control:
            self.control_pressed = True
        if self.control_pressed:
            if e.key() == Qt.Key_Q:
                self.control_preconditions()
            if e.key() == Qt.Key_W:
                self.control_actual_results()
            if e.key() == Qt.Key_1:
                self.updateTestStatus("PASS")
            if e.key() == Qt.Key_2:
                self.updateTestStatus("FAIL")
            if e.key() == Qt.Key_3:
                self.updateTestStatus("BLOCKED")
            if e.key() == Qt.Key_4:
                self.updateTestStatus("NOT TESTED")
            return
        if e.key() == Qt.Key_Q:
            self.previousTestCase()
        if e.key() == Qt.Key_W:
            self.nextTestCase()


    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragging = True
            self.drag_position = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event):
        if self.dragging:
            self.move(event.globalPos() - self.drag_position)
            event.accept()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragging = False
            event.accept()

    def handleExport(self):
        selected_option = self.export_combo.currentText()
        match selected_option.lower():
            case "jira":
                report = self._prepare_jira_export()
            case "xlsx":
                report = self._prepare_xlsx_export()
            case "export":
                return
            case _:
                QMessageBox.critical(self, self.app_name, f"'{selected_option}' not supported as an export!")
                return
        #raise NotImplemented
        self.export_combo.setCurrentIndex(0)
        report_box = QMessageBox()
        report_box.setGeometry(100,100,500,300)
        report_box.setWindowTitle(f"{selected_option} report")
        report_box.setText(report[0])
        report_box.setDetailedText(report[1])
        report_box.setWindowFlag(Qt.WindowStaysOnTopHint)
        report_box.setMinimumWidth(500)
        report_box.setStyleSheet("""QLabel{
        min-width: 300px;
        min-height: 30px;
        },
        QTextEdit{
        min-height: 450px;
        }
        """)
        report_box.exec()

    def _prepare_jira_export(self) -> tuple:
        f = Formatting("jira")
        passed = []
        failed = []
        blocked = []
        not_tested = []
        for tc in self.test_cases:
            match str(tc.test_status).upper():
                case "PASS":
                    passed.append(tc)
                case "FAIL":
                    failed.append(tc)
                case "BLOCKED":
                    blocked.append(tc)
                case "NOT TESTED":
                    not_tested.append(tc)
                case _:
                    not_tested.append(tc)
        report = f"""
| PASSED | {len(passed)} |
| FAILED | {len(failed)} |
| BLOCKED | {len(blocked)} |
| NOT TESTED | {len(not_tested)} |

"""
        report += f"\n{f.f['h2']}Failed tests\n"
        for tc in failed:
            report += f"{f.f['bullet']} {tc}\n"
        report += f"\n{f.f['h2']}Blocked tests\n"
        for tc in blocked:
            report += f"{f.f['bullet']} {tc}\n"
        report += f"\n{f.f['h2']}Passed tests\n"
        for tc in passed:
            report += f"{f.f['bullet']} {tc}\n"
        report += f"\n{f.f['h2']}Not tested tests\n"
        for tc in not_tested:
            report += f"{f.f['bullet']} {tc}\n"
        return ("Jira export is ready!\nClick \"Show Details\"", report)

    def _prepare_xlsx_export(self) -> tuple:
        columns = ["A","B","C","D","E","F","G","H","I","J","K","L"]
        try:
            from openpyxl import Workbook
        except ImportError:
            err_msg = "Could not load openpyxl!\nInstall openpyxl or import json file!"
            QMessageBox.critical(self, title="PyTestCases Error: Could not export Xlsx", text=err_msg)
            raise ImportError(err_msg)
        # save results before export
        self.saveResults()
        export_file = f"export_{self.test_execution_id}.xlsx"
        wb = Workbook()
        ws = wb.active
        for _ in range(self.test_case_data_starting_row-2):
            ws.append([])
        ws.append(["Area", "Test Case Id", "Feature", "Label", "Level", "Test Case Name", "Test Steps", "Preconditions", "Test Step Description", "Expected Result", "Test Status", "Assignee", "Actual Result", "Date Executed", "Notes"])
        ws.title = f"{self.test_execution_id}"
        report_msg = f"Following test cases were exported to {export_file}:"
        report = ""
        row_i = 0
        for i, tc in enumerate(self.test_cases):
            report += str(tc) + "\n"
            tc_data = tc.dict_to_save()
            #for field_name, field_values in tc_data.items():
            if True:
                for steps_i in range(len(tc_data["Test Steps"])):
                    row_export_data = []
                    row_export_data.append(tc_data["Area"])
                    row_export_data.append(tc_data["Test Case Id"])
                    row_export_data.append("")#tc_data["Feature"])
                    row_export_data.append("")#tc_data["Label"])
                    row_export_data.append(tc_data["Level"])
                    row_export_data.append(tc_data["Test Case Name"])
                    row_export_data.append("")#tc_data["Test Steps"]
                    row_export_data.append(tc_data["Preconditions"])
                    tc_data["Preconditions"] = ""
                    row_export_data.append(tc_data["Test Steps"][steps_i][0]) # Test Step Description
                    row_export_data.append(tc_data["Test Steps"][steps_i][1]) # Expected Result
                    row_export_data.append(tc_data["Test Status"])
                    row_export_data.append(tc_data["Assignee"])
                    row_export_data.append(tc_data["Test Steps"][steps_i][2])# Actual Result
                    row_export_data.append("")#tc_data["Date Executed"]
                    row_export_data.append(tc_data["Test Steps"][steps_i][3]) # Notes
                    ws.append(row_export_data)
                continue
        wb.save(export_file)
        return (report_msg, report)


    def displayTestCase(self):
        self.current_test_index = self.test_case_dropdown.currentIndex()
        if self.current_test_index == -1:
            return

        test_case = self.test_cases[self.current_test_index].dict()
        self.setTestStatus(test_case["Test Status"])

        self.test_case_table.setRowCount(0)
        for step, expected, actual, note in zip(
                test_case["Test Steps"][0],
                test_case["Test Steps"][1],
                test_case["Test Steps"][2],
                test_case["Test Steps"][3],
        ):
            row_position = self.test_case_table.rowCount()
            self.test_case_table.insertRow(row_position)
            self.test_case_table.setItem(row_position, 0, QTableWidgetItem(step))
            self.test_case_table.setItem(row_position, 1, QTableWidgetItem(expected))
            self.test_case_table.setItem(row_position, 2, QTableWidgetItem(actual))
            self.test_case_table.resizeRowsToContents()

        preconditions = test_case.get("Preconditions", "")
        self.preconditions_text.clear()
        if preconditions:
            self.preconditions_label.show()
            self.preconditions_text.show()
            self.preconditions_text.setText(preconditions)
        else:
            self.preconditions_label.hide()
            self.preconditions_text.hide()



        if test_case.get("Assignee"):
            self.assignee_label.show()
            self.assignee_value.setText(test_case["Assignee"])
            self.assignee_value.show()

    def setTestStatus(self, test_status: str):
        status = str(test_status).upper()
        color_map = {
            "PASS": "green",
            "FAIL": "red",
            "BLOCKED": "yellow",
            "NOT TESTED": "gray"
        }
        color = color_map.get(status, "gray")
        self.test_status_label.setText(f"<b style='color:{color};'>{status}</b>")

    def loadTests(self, event=None, file_path_arg:str|None = None):
        options = QFileDialog.Options()
        if file_path_arg is None:
            file_path, _ = QFileDialog.getOpenFileName(self, "Load Test File", "", "All Files (*);;Excel Files (*.xlsx);;JSON Files (*.json)", options=options)
        else:
            file_path = file_path_arg
        if not file_path:
            return

        self.file_name = Path(file_path).name

        if file_path.endswith("xlsx"):
            raw_data = self.loadXlsx(file_path)
            if raw_data is None:
                return
        elif file_path.endswith("json"):
            raw_data = self.loadJson(file_path)
        else:
            QMessageBox.critical(self, "Error", "File format not supported")
            return

        self.loadTestsFromRawData(raw_data)

        if raw_data[0] == "from_output":
            execution_id = self.file_name.removeprefix(self.default_exports_prefix).removesuffix(".json")
            self.execution_id_input.setText(execution_id)
            self._update_test_execution_id()
            self.file_loaded_from_output = True
            self.execution_id_input.setDisabled(True)
        else:
            self.file_loaded_from_output = False
            current_text_execution = self.execution_id_input.text()
            if not current_text_execution:
                self.test_execution_id = str(int(time()))
                self.execution_id_input.setText(self.test_execution_id)

        self.test_case_dropdown.clear()
        self.test_case_dropdown.addItems([str(test_case) for test_case in self.test_cases])

        self.current_test_index = 0
        self.test_case_dropdown.setCurrentIndex(self.current_test_index)
        self.displayTestCase()

    def updateTestStatus(self, status: str):
        if not len(self.test_cases):
            return
        self.test_cases[self.current_test_index].set_status(status)
        self.setTestStatus(status)
        self.saveResults()

    def nextTestCase(self):
        if not len(self.test_cases):
            return
        if self.current_test_index < len(self.test_cases) - 1:
            self.current_test_index += 1
            self.test_case_dropdown.setCurrentIndex(self.current_test_index)
            self.displayTestCase()

    def previousTestCase(self):
        if not len(self.test_cases):
            return
        if self.current_test_index > 0:
            self.current_test_index -= 1
            self.test_case_dropdown.setCurrentIndex(self.current_test_index)
            self.displayTestCase()

    def _update_test_execution_id(self):
        self.test_execution_id = self.execution_id_input.text()

    def _return_test_steps(self):
        test_steps = [[],[],[],[]]
        row_i, column_i = 0, 0
        for column in range(self.test_case_table.columnCount()):
            for row in range(self.test_case_table.rowCount()):
                _item = self.test_case_table.item(row, column)
                if _item:
                    item = self.test_case_table.item(row, column).text()
                    test_steps[column].append(item)
        return test_steps

    def saveResults(self):
        self._update_test_execution_id()
        self.test_execution_id = self.execution_id_input.text()

        if not self.file_loaded_from_output and not self.test_execution_id:
            QMessageBox.warning(self, "Error", "Test Execution ID is required!")
            return

        if self.file_loaded_from_output:
            output_file_name = self.file_name
        else:
            output_file_name = f"{self.default_exports_prefix}{self.test_execution_id}.json"

        # read data from tables
        if not self.editing_disabled:
            data_from_table = self._return_test_steps()
            self.test_cases[self.current_test_index].set_test_steps(data_from_table)
        #if isinstance(self.test_cases, str):
        #    test_case_data_to_save.insert(0, self.test_cases[0])
        test_case_data_to_save = [tc.dict_to_save() for tc in self.test_cases if isinstance(tc, TestCaseModel)]
        test_case_data_to_save.insert(0, "from_output")
        with open(output_file_name, "w") as file:
            json.dump(test_case_data_to_save, file, indent=4)

        # inform user that it worked
        if self.show_save_info:
            QMessageBox.information(self, "Info", f"This pop-up is shown once per session!\nResults saved to {output_file_name}")
            self.show_save_info = False
        #self.test_cases.remove("from_output")


    def loadJson(self, file_path: str):
        with open(file_path) as f:
            return json.load(f)

    def loadXlsx(self, file_path: str):
        try:
            from openpyxl import load_workbook
        except ImportError:
            err_msg = "Could not load openpyxl!\nInstall openpyxl or import json file!"
            QMessageBox.critical(self, title="PyTestCases Error: Could not load Xlsx", text=err_msg)
            raise ImportError(err_msg)
        wb = load_workbook(file_path, data_only=True)
        try:
            ws = wb[self.xlsx_test_cases_sheet_name]
        except KeyError:
            # TODO it's messy, but it does the job for now
            win = QMainWindow()
            win_widg = QWidget(win)
            win.setCentralWidget(win_widg)
            win.setWindowTitle(self.app_name)
            win_layout = QVBoxLayout(win_widg)
            win_label = QLabel("Enter Sheet name", win_widg)
            win_entry = QLineEdit(win)
            win_ok = QPushButton("Ok", win)
            win_ok.clicked.connect(lambda: [setattr(self, "xlsx_test_cases_sheet_name", win_entry.text()), win.destroy(), self.loadTests(file_path_arg=file_path)])
            win_detect = QPushButton("Detect Sheets", win_widg)
            win_detect.clicked.connect(lambda: [self._detect_sheet_names(file_path)])
            win_layout.addWidget(win_label)
            win_layout.addWidget(win_entry)
            win_layout.addWidget(win_ok)
            win_layout.addWidget(win_detect)
            win.show()
            return 
            #QMessageBox.critical(self, self.app_name, f"Did not detect sheet with name '{self.xlsx_test_cases_sheet_name}'!")
            #return

        ws = wb[self.xlsx_test_cases_sheet_name]
        column_names = [cell.value for cell in ws[self.column_names_row]]

        raw_data = []
        current_test_case = None
        current_test_steps = []
        current_test_preconditions = ""

        # get the hidden rows
        hidden_rows = set()
        for row in ws.row_dimensions:
            if ws.row_dimensions[row].hidden:
                hidden_rows.add(row)

        # gather data
        for row in ws.iter_rows(min_row=self.test_case_data_starting_row, values_only=False):
            # continue if row is hidden
            if row[0].row in hidden_rows:
                continue
            # get data if row is visible
            row_data = dict(zip(column_names, [c.value for c in row]))

            if current_test_case and row_data["Test Case Id"] != current_test_case["Test Case Id"]:
                tc = {
                    "Test Case Id": int(float(current_test_case["Test Case Id"])),
                    "Test Case Name": current_test_case["Test Case Name"],
                    "Area": current_test_case["Area"],
                    "Level": current_test_case["Level"],
                    "Preconditions": current_test_preconditions,
                    "Test Steps": current_test_steps,
                    "Test Execution Id": current_test_case.get("Test Execution Id"),
                    "Test Status": current_test_case.get("Test Status"),
                    "Assignee": current_test_case.get("Assignee"),
                    "Gsheet Document Id": current_test_case.get("Gsheet Document Id")
                }
                raw_data.append(tc)
                current_test_steps = []
                current_test_preconditions = ""

            current_test_case = row_data
            current_test_steps.append([row_data["Test Step Description"], row_data["Expected Result"], row_data["Actual Result"], row_data["Notes"]])
            current_test_preconditions += (str(row_data["Preconditions"]) + "\n") if row_data.get("Preconditions") is not None else ""


        if current_test_case:
            tc = {
                    "Test Case Id": current_test_case["Test Case Id"],
                    "Test Case Name": current_test_case["Test Case Name"],
                    "Area": current_test_case["Area"],
                    "Level": current_test_case["Level"],
                    "Preconditions": current_test_preconditions,
                    "Test Steps": current_test_steps,
                    "Test Execution Id": current_test_case.get("Test Execution Id"),
                    "Test Status": current_test_case.get("Test Status"),
                    "Assignee": current_test_case.get("Assignee"),
                    "Gsheet Document Id": current_test_case.get("Gsheet Document Id")
                }
            raw_data.append(tc)

        while raw_data[-1] is None:
            raw_data.pop()
        while raw_data[-1]["Test Case Id"] is None:
            raw_data.pop()
        return raw_data
        self.test_cases += raw_data


    def control_preconditions(self):
        if self.preconditions_displayed:
            self.preconditions_text_frame.hide()
            self.preconditions_displayed = False
        else:
            self.preconditions_text_frame.show()
            self.preconditions_displayed = True


    def control_actual_results(self):
        #print(self.geometry())
        # TODO resize window
        if not self.actual_results_displayed:
            self.hide_actual_results()
            self.actual_results_displayed = True
            self.test_case_table_control_actual_resultss_btn.setText(">")
            self.setFixedWidth(int(self.column_width*2))
        else:
            self.show_actual_results()
            self.actual_results_displayed = False
            self.test_case_table_control_actual_resultss_btn.setText("<")
            self.setFixedWidth(int(self.column_width*3))
        window_width = int(self.geometry().width())
        self.test_case_table_control_actual_resultss_btn.move(window_width-45,10)

    def hide_actual_results(self):
        self.test_case_table.setColumnHidden(2, True)

    def show_actual_results(self):
        self.test_case_table.setColumnHidden(2, False)

######
    def loadTestsFromRawData(self, raw_data: list[dict]):
        for i, test_case in enumerate(raw_data):
            if isinstance(test_case, str):
                if test_case == "from_output":
                    continue
            try:
                tc = TestCaseModel(
                        test_case_id=test_case["Test Case Id"],
                        test_case_name=test_case["Test Case Name"],
                        area=test_case["Area"],
                        level=test_case["Level"],
                        test_steps=test_case["Test Steps"],
                        #expected_results=test_case["Test Steps"][1],
                        #actual_results=test_case["Test Steps"][2] if len(test_case["Test Steps"]) > 2 else None,
                        #notes=test_case["Test Steps"][3] if len(test_case["Test Steps"]) > 3 else None,
                        test_status=test_case.get("Test Status"),
                        preconditions=test_case.get("Preconditions"), 
                        assignee=test_case.get("Assignee"),
                        gsheet_document_id=test_case.get("Gsheet Document Id")
                        )
                self.test_cases.append(tc)
            except TypeError as e:
                print(test_case)
                QMessageBox.warning(self, self.app_name, f"Could not convert '{test_case}' to Test Case due to TypeError: {e}")
            except KeyError as e:
                print(test_case)
                QMessageBox.error(self, self.app_name, f"Could not load test case with index {i} due to KeyError: {e}!")

    def _detect_sheet_names(self, file_path):
        if "load_workbook" not in locals():
            from openpyxl import load_workbook
        sheet_names = load_workbook(file_path).sheetnames
        sheet_names_as_str = f"Detected Sheet names for file {file_path}\n\n"
        for sn in sheet_names:
            sheet_names_as_str += " - " + sn + "\n"
        QMessageBox.information(self, self.app_name, sheet_names_as_str)

        


if __name__ == "__main__":
    app = QApplication([])
    window = PyTestCasesApp()
    app.exec_()


